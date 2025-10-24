import requests
from datetime import date, datetime, timezone, timedelta
from typing import List
import sys
import json
import pandas as pd
from zoneinfo import ZoneInfo
import openpyxl
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import uuid
from lxml import etree
from pathlib import Path
import base64

class Trader:
    def __init__(self, username: str, password: str, base_url: str, flow_date: date):
        """
        Initialize the Trader and perform login.

        :param username: API username
        :param password: API password
        :param base_url: API base URL
        """
        self.username = username
        self.password = password
        self.base_url = base_url
        self.session = requests.Session()
        self.token = None
        self.rome_tz = ZoneInfo("Europe/Rome")

        flow_date = datetime.combine(flow_date, datetime.min.time(), tzinfo=self.rome_tz)
        flow_date = flow_date.astimezone(ZoneInfo("UTC"))
        next_flow_date = flow_date + timedelta(days=1)
        self.flow_date = flow_date.strftime("%Y-%m-%dT%H:%M:%SZ")
        self.next_flow_date = next_flow_date.strftime("%Y-%m-%dT%H:%M:%SZ")

        # Perform login immediately
        self.market_player = self.login()

    def login(self):
        """
        Perform login and store token/session.
        Adapt this function to your specific API requirements.
        """
        response = self.session.post(self.base_url + '/login', json={'username': self.username, 'password': self.password})
        response.raise_for_status()
        self.token = response.json()['token']
        self.session.headers['Authorization'] = f'Bearer {self.token}'
        print("Login successful.")
        return response.json()['user']['_id']

    def fetch_auction(self):

        query = {
            'resolution': 'PT15M',
            'date': self.flow_date,
            'market': 'MI1',
            'status': 'Accept'
            # or 'PT15M' depending on your use case
        }

        response = self.session.request(
            'get',
            f"{self.base_url}/offers",
            params={
                'query$': json.dumps(query),
            }
        )

        orders = response.json().get('data', [])
        orders = pd.DataFrame(orders)
        orders = orders.loc[:, ['awarded_price', 'awarded_quantity', 'delivery_start', 'unit_reference_number']]

        orders["delivery_start"] = pd.to_datetime(orders["delivery_start"], utc=True)
        orders["delivery_start"] = orders["delivery_start"].dt.tz_convert("Europe/Rome")
        orders["period"] = orders["delivery_start"].dt.hour * 4 + orders["delivery_start"].dt.minute // 15 + 1
        orders["flow_date"] = orders["delivery_start"].dt.date

        orders = orders.drop(columns='delivery_start')
        orders.loc[:, 'unit_reference_number'] = orders.loc[:, 'unit_reference_number'].apply(lambda z: z.replace('UC_DP2502_',''))
        orders = orders.groupby(['awarded_price', 'unit_reference_number', 'period', 'flow_date']).sum().reset_index()

        file = r"C:\Users\lbellomi\PycharmProjects\pythonProject\Trader\bids.xlsm"
        wb = openpyxl.load_workbook(file, keep_vba=True)

        prices = pd.read_excel(r'C:\Users\lbellomi\PycharmProjects\pythonProject\Trader\bids.xlsm', sheet_name=None)
        prices = prices['Prices']
        prices = prices.iloc[:, [0, 14, 15, 16, 17, 21]]
        prices = prices.rename(columns={'NORD.1': 'NORD', 'SICI.1': 'SICI', 'SUD.1': 'SUD', 'CSUD.1': 'CSUD', 'Unnamed: 21': 'date'})
        prices.loc[:, 'flow_date'] = prices.loc[:, 'date'].apply(lambda d: d.date())
        prices = prices.melt(id_vars=["period", 'flow_date'], var_name="unit_reference_number", value_name="MGP")

        orders = orders.merge(prices.loc[:, ['period','flow_date','unit_reference_number', 'MGP']], on=['period','flow_date','unit_reference_number'], how='left')
        orders.loc[:, 'diff'] = orders.loc[:, 'MGP'] - orders.loc[:, 'awarded_price']

        for ind in orders.index:

            if orders.loc[ind, 'diff'] <= 14.9:
                orders.loc[ind, 'perc'] = 0.2999

            elif (orders.loc[ind, 'diff'] > 14.9) and (orders.loc[ind, 'diff'] < 20):
                orders.loc[ind, 'perc'] = 0.35

            else:
                orders.loc[ind, 'perc'] = 0.4

        orders.loc[:, 'bids'] = orders.loc[:, 'awarded_price'] + orders.loc[:, 'diff'] * orders.loc[:, 'perc']

        bids = wb["MI1"]
        start_row = 2

        for i, (price, period, zone, qty) in enumerate(zip(orders["bids"], orders["period"], orders["unit_reference_number"], orders["awarded_quantity"]), start=start_row):

            bids.cell(row=i, column=3, value=price)
            bids.cell(row=i, column=2, value=period)
            bids.cell(row=i, column=6, value=zone)
            bids.cell(row=i, column=4, value=qty)
            bids.cell(row=i, column=1, value='SELL')
            bids.cell(row=i, column=5, value='PT15M')

        wb.save(file)

    def place_orders(self, zone: list, granularity: str, purpose_list: list, period: list, price: list, qty: list, message: str  = ''):
        """
        adds orders on the specific book
        """
        if (len(period) != len(price)) or (len(price)!= len(qty)):
            print('Dimensione prezzi e quantità discordanti')
            sys.exit()

        dim = len(period)

        unit_code = [f"UC_DP2502_{z}" for z in zone]

        payload = create_payload(pos_list=period, purpose_list=purpose_list, price_list=price, qty_list=qty,  area_code=zone, unit_code=unit_code, flow_date=self.flow_date, granularity=granularity, message=message)

        url = f"{self.base_url}/xbid/books/orders"
        response = self.session.post(url, json=payload)
        if response.status_code == 200:
            print("✅ Order posted successfully.")
            print(response.json())
        else:
            raise Exception(f"❌ Error {response.status_code}: {response.text}")

    def imbalance_position(self, unit_id: List[str]):

        df = pd.DataFrame()

        for id in unit_id:

            response = self.session.request(
                'get',
                self.base_url + '/units/programs-and-economics',
                params={
                    'delivery_from': self.flow_date,
                    'delivery_to': self.next_flow_date,
                    'unit': 'UC_DP2502_' + id,
                })

            data = pd.DataFrame(response.json()['data'])

            if not data.empty:
                data.loc[:, 'zone'] = id
                df = pd.concat([df, data])

        df = df.loc[:, ['commercial_imbalance','delivery_start','qty_MGP','qty_MI1', 'qty_XBID', 'qty_MI2', 'zone']]

        df["delivery_start"] = pd.to_datetime(df["delivery_start"], utc=True)
        df["delivery_start"] = df["delivery_start"].dt.tz_convert("Europe/Rome")
        df["period"] = df["delivery_start"].dt.hour * 4 + df["delivery_start"].dt.minute // 15 + 1
        df["flow_date"] = df["delivery_start"].dt.date

        df = df.loc[:, ['flow_date','period', 'zone', 'qty_MGP','qty_MI1', 'qty_XBID', 'qty_MI2', 'commercial_imbalance']]
        df = df.loc[df.loc[:, 'commercial_imbalance'] != 0, :]

        return df

    def generate_bids(self):

        prices = pd.read_excel(r'C:\Users\lbellomi\PycharmProjects\pythonProject\Trader\bids.xlsm', sheet_name=None)
        prices = prices['Prices']

        prices = prices.iloc[:, [0, 1, 2, 3, 4, 14, 15, 16, 17, 21]]
        prices = prices.rename(columns={'NORD.1': 'NORD_MGP', 'SICI.1': 'SICI_MGP', 'SUD.1': 'SUD_MGP', 'CSUD.1': 'CSUD_MGP', 'NORD': 'NORD_MI1', 'SICI': 'SICI_MI1', 'SUD': 'SUD_MI1', 'CSUD': 'CSUD_MI1', 'Unnamed: 21': 'date'})
        prices.loc[:, 'flow_date'] = prices.loc[:, 'date'].apply(lambda d: d.date())

        prices = prices.drop(columns='date')

        prices.loc[:, 'DIFF_NORD'] = prices.loc[:, 'NORD_MI1'] - prices.loc[:, 'NORD_MGP']
        prices.loc[:, 'DIFF_SUD'] = prices.loc[:, 'SUD_MI1'] - prices.loc[:, 'SUD_MGP']
        prices.loc[:, 'DIFF_SICI'] = prices.loc[:, 'SICI_MI1'] - prices.loc[:, 'SICI_MGP']
        prices.loc[:, 'DIFF_CSUD'] = prices.loc[:, 'CSUD_MI1'] - prices.loc[:, 'CSUD_MGP']

        prices = prices.loc[:, ['period', 'flow_date', 'NORD_MI1', 'CSUD_MI1', 'SUD_MI1', 'SICI_MI1', 'DIFF_NORD', 'DIFF_SUD', 'DIFF_CSUD', 'DIFF_SICI']]

        bids = pd.DataFrame(columns=['PURPOSE','PERIOD', 'PRICE', 'QTY','TIME', 'ZONA'])

        Pr_std = {'NORD': [8.5, 13, 19.5], 'SUD': [10, 13, 19.5], 'CSUD': [8.5, 13, 19.5], 'SICI': [10, 13, 19.5]}
        Pr_ex = {'NORD': [1.75, -3], 'SUD': [1.75, -3], 'CSUD': [1.75, -3], 'SICI': [1.75, -3]}

        Qt_std = {'NORD': [3, 9, 20], 'SUD': [3, 9, 20], 'CSUD': [3, 9, 20], 'SICI': [2, 6, 12]}
        Qt_ex = {'NORD': [12, 20], 'SUD': [12, 20], 'CSUD': [12, 20], 'SICI': [8, 12]}

        for zona in ['NORD', 'SUD', 'SICI', 'CSUD']:

            temp = prices.loc[:, ['period', 'flow_date', zona + '_MI1', 'DIFF_' + zona]]
            temp_norm = temp.loc[temp.loc[:, 'DIFF_' + zona] < 15, :]
            temp_extra = temp.loc[temp.loc[:, 'DIFF_' + zona] >= 15, :]

            for i in zip(Pr_std[zona], Qt_std[zona]):

                temp_b = pd.DataFrame(columns=['PURPOSE','PERIOD', 'PRICE','QTY','TIME', 'ZONA'])
                temp_b.loc[:, 'PERIOD'] = temp_norm.loc[:, 'period'].values
                temp_b.loc[:, 'PURPOSE'] = 'BUY'
                temp_b.loc[:, 'QTY'] = i[1]
                temp_b.loc[:, 'PRICE'] = (temp_norm.loc[:, zona + '_MI1'] - i[0]).values
                temp_b.loc[:, 'TIME'] = 'PT15M'
                temp_b.loc[:, 'ZONA'] = zona
                bids = pd.concat([bids, temp_b])

            if temp_extra.shape[0] == 1:

                for i in zip(Pr_ex[zona], Qt_ex[zona]):
                    temp_b = pd.DataFrame(columns=['PURPOSE', 'PERIOD', 'PRICE', 'QTY', 'TIME'])
                    temp_b.loc[0, 'PERIOD'] = temp_extra.loc[:, 'period'].values[0]
                    temp_b.loc[0, 'PURPOSE'] = 'BUY'
                    temp_b.loc[0, 'QTY'] = i[1]
                    temp_b.loc[0, 'PRICE'] = (temp_extra.loc[:, zona + '_MI1'] - temp_extra.loc[:, 'DIFF_' + zona] + i[0]).values[0]
                    temp_b.loc[0, 'TIME'] = 'PT15M'
                    temp_b.loc[0, 'ZONA'] = zona
                    bids = pd.concat([bids, temp_b])

            elif not temp_extra.empty:

                for i in zip(Pr_ex[zona], Qt_ex[zona]):
                    temp_b = pd.DataFrame(columns=['PURPOSE', 'PERIOD', 'PRICE', 'QTY', 'TIME'])
                    temp_b.loc[:, 'PERIOD'] = temp_extra.loc[:, 'period'].values
                    temp_b.loc[:, 'PURPOSE'] = 'BUY'
                    temp_b.loc[:, 'QTY'] = i[1]
                    temp_b.loc[:, 'PRICE'] = (temp_extra.loc[:, zona + '_MI1'] - temp_extra.loc[:, 'DIFF_' + zona] + i[0]).values
                    temp_b.loc[:, 'TIME'] = 'PT15M'
                    temp_b.loc[:, 'ZONA'] = zona
                    bids = pd.concat([bids, temp_b])
            else:
                continue

        bids = bids.reset_index(drop=True)

        file_path = r'C:\Users\lbellomi\PycharmProjects\pythonProject\Trader\bids.xlsm'
        sheet_name = "BIDS"
        rows_per_block = 100
        cols_per_block = 6
        col_gap = 1

        # Load workbook and sheet
        wb = load_workbook(file_path, keep_vba=True)
        ws = wb[sheet_name]

        # Optional: clear previous content
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

        # Number of blocks (each block = up to 100 rows)
        num_blocks = -(-len(bids) // rows_per_block)  # ceiling division

        base_headers = ["PURPOSE", "PERIOD", "PRICE", "QTY", "TIME", "ZONA"]

        for block_index in range(num_blocks):
            start_row = 2
            start_col = block_index * (cols_per_block + col_gap) + 1

            headers = [f"{name}-{block_index + 1}" for name in base_headers]

            # --- Write headers ---
            for c_idx, header in enumerate(headers, start=start_col):
                ws.cell(row=1, column=c_idx, value=header)

            # Select slice of data
            block_df = bids.iloc[
                       block_index * rows_per_block: (block_index + 1) * rows_per_block
                       ]

            # Write each block
            for r_idx, row in enumerate(block_df.itertuples(index=False), start=start_row):
                for c_idx, value in enumerate(row, start=start_col):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        # Save workbook
        wb.save(file_path)

    def bid_auction(self, pos_list, purpose_list, price_list, qty_list, area_code, unit_code, granularity, market):

        url = f"{self.base_url}/market/send-offers"
        payload = create_auction_payload(pos_list, purpose_list, price_list, qty_list, area_code, unit_code, self.flow_date, granularity, market)

        response = self.session.post(url, json=payload)
        if response.status_code == 200:
            print("✅ Order posted successfully.")
            print(response.json())
        else:
            raise Exception(f"❌ Error {response.status_code}: {response.text}")

    def build_xbid_orders_xml(self,
            zone, granularity, purpose_list, period, price, qty,
            operator_code="110961",
            unit_id="UC_DP2502_",
            receiver_operator="IDGME",
            execution="None",
            expiry_time=None
    ):

        """
        Build an XML-LTS message for XBID orders from parallel input lists.

        Parameters
        ----------
        zone : list[str]
            Bidding zone (e.g. ["NORD", "CNOR"])
        granularity : list[str]
            Granularity strings, "PT15M" (quarter-hour) or "PT60M" (hour)
        purpose_list : list[str]
            "BUY" or "SELL" (converted to 'P' or 'S' internally)
        period : list[int]
            Delivery intervals (1–96 for QH, 1–24 for hourly)
        price : list[float]
            Offer price in €/MWh
        qty : list[float]
            Offer qty in MW
        flow_date : str, optional
            Delivery date (YYYY-MM-DD). Defaults to today's date.
        operator_code, unit_id, receiver_operator, execution, expiry_time : str
            Extra parameters for GME XML header and fields.

        Returns
        -------
        bytes : XML document encoded as iso-8859-1
        """

        # --- Input validation ---
        n = len(zone)
        if not all(len(lst) == n for lst in [granularity, purpose_list, period, price, qty]):
            raise ValueError("All input lists must have the same length")

        # Default dates
        now = datetime.now(timezone.utc)
        msg_date = now.strftime("%Y-%m-%d")
        msg_time = now.strftime("%H:%M:%S.%fZ")
        if self.flow_date is None:
            flow_date = msg_date
        if expiry_time is None:
            # default expiry at 23:00 UTC same day
            expiry_time = self.flow_date

        # --- XML namespaces ---

        root = etree.Element(
            "Message",
            xmlns="urn:XML-LTS",
            MessageType="Request",
            MessageDate=msg_date,
            MessageTime=msg_time
        )

        # --- Header ---
        header = etree.SubElement(root, "Header")
        sender = etree.SubElement(header, "Sender")
        etree.SubElement(sender, "CompanyName").text = operator_code
        etree.SubElement(sender, "UserMsgCode").text = "user"
        etree.SubElement(sender, "OperatorMsgCode").text = operator_code

        receiver = etree.SubElement(header, "Receiver")
        etree.SubElement(receiver, "OperatorMsgCode").text = receiver_operator

        # --- Transaction ---
        transaction = etree.SubElement(root, "Transaction")
        basket = etree.SubElement(transaction, "OffersBasket")
        etree.SubElement(basket, "Execution").text = execution
        offers_outer = etree.SubElement(basket, "Offers")

        count = 0

        # --- Offers loop ---
        for zone, gran, purpose, period, price, qty in zip(zone, granularity, purpose_list, period, price, qty):
            count += 1
            offer_el = etree.SubElement(offers_outer, "Offers")
            etree.SubElement(offer_el, "OperatorCode").text = operator_code
            etree.SubElement(offer_el, "FlowDate").text = self.flow_date
            etree.SubElement(offer_el, "ZoneCode").text = zone
            etree.SubElement(offer_el, "UnitId").text = unit_id + zone
            etree.SubElement(offer_el, "ExternalNotes").text = 'ORD_' + str(count)

            # Determine interval type
            interval_type = "QH" if gran == "PT15M" else "FH"
            interval_el = etree.SubElement(offer_el, "Interval", type=interval_type)
            interval_el.text = str(period)

            # Purpose: convert BUY/SELL to P/S
            etree.SubElement(offer_el, "Purpose").text = "P" if purpose.upper() == "BUY" else "S"
            etree.SubElement(offer_el, "Status").text = "A"
            etree.SubElement(offer_el, "ExpiryTime").text = expiry_time
            etree.SubElement(offer_el, "Qty").text = str(qty)
            etree.SubElement(offer_el, "Price").text = str(price)

        # --- Serialize ---
        xml_bytes = etree.tostring(
            root,
            pretty_print=True,
            encoding="iso-8859-1",
            xml_declaration=True
        )

        output_path = Path(r'C:\Users\lbellomi\PycharmProjects\pythonProject\Trader\orders\orders_' + self.flow_date.replace(":", "-") + '_XBID.xml')
        output_path.write_bytes(xml_bytes)

    def submit_xbid_xml(self):
        # Read file
        xml_path = r'C:\Users\lbellomi\PycharmProjects\pythonProject\Trader\orders\orders_' + self.flow_date.replace(":", "-") + '_XBID.xml'
        xml_bytes = Path(xml_path).read_bytes()
        # Encode to base64
        xml_b64 = base64.b64encode(xml_bytes).decode("ascii")

        # Prepare JSON payload
        payload = {
            "name": Path(xml_path).name,
            "content": xml_b64
        }

        headers = {
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json"
        }

        # POST to Hermes
        response = requests.post(self.base_url + '/files/xbidws', json=payload, headers=headers)
        if response.status_code == 200:
            print("✅ File successfully submitted to Hermes.")
        else:
            print("❌ Submission failed:", response.status_code, response.text)

        return response.json()

    def compute_position(self):
        """
        Computes the current position, with size exposure and weighted avg price
        """

        # fetching the trades and rearranging them
        query = {
            'resolution': 'PT15M',  # or 'PT15M' depending on your use case
        }

        response = self.session.request(
            'get',
            f"{self.base_url}/trades",
            params={
                'delivery_from': self.flow_date,
                'delivery_to': self.next_flow_date,
                'query$': json.dumps(query),
            }
        )

        trades = response.json().get('data', [])
        trades = pd.DataFrame(trades)
        trades = trades.loc[:,
                 ['buyer_unit_code', 'delivery_start', 'price', 'quantity', 'resolution', 'seller_unit_code', 'buyer_hermes_txt', 'seller_hermes_txt']]

        trades.loc[trades.loc[:, 'buyer_unit_code'].isna() == False, 'zona'] = trades.loc[
            trades.loc[:, 'buyer_unit_code'].isna() == False, 'buyer_unit_code']
        trades.loc[trades.loc[:, 'seller_unit_code'].isna() == False, 'zona'] = trades.loc[
            trades.loc[:, 'seller_unit_code'].isna() == False, 'seller_unit_code']
        trades.loc[:, 'zona'] = trades.loc[:, 'zona'].apply(lambda d: d.split('_')[2])
        trades.loc[trades.loc[:, 'buyer_unit_code'].isna() == False, 'type'] = 'BUY'
        trades.loc[trades.loc[:, 'seller_unit_code'].isna() == False, 'type'] = 'SELL'
        trades = trades.drop(columns=['buyer_unit_code', 'seller_unit_code'])
        trades.loc[:, 'delivery_start'] = trades.loc[:, 'delivery_start'].apply(
            lambda d: datetime.strptime(d, "%Y-%m-%dT%H:%M:%S").replace(tzinfo=ZoneInfo("UTC")))
        trades.loc[:, 'delivery_start'] = trades.loc[:, 'delivery_start'].apply(lambda d: d.astimezone(self.rome_tz))
        trades.loc[:, 'period'] = trades.loc[:, 'delivery_start'].apply(lambda t: t.hour * 4 + t.minute // 15 + 1)

        unit_codes = list(trades.loc[:, 'zona'].unique())
        position = self.imbalance_position(unit_codes)

        order_book = pd.DataFrame()

        for resolution in trades.loc[:, 'resolution'].unique():
            for area_code in trades.loc[:, 'zona'].unique():

                # fetching the books and rearranging them
                params = {
                    'flowDate': self.flow_date,
                    'resolution': resolution,
                    'deliveryAreaId': area_code,
                    'marketPlayer': self.market_player
                }

                response = self.session.get(f"{self.base_url}/xbid/books", params=params)

                book = response.json()["data"]
                book = pd.DataFrame(book)

                books = pd.DataFrame(columns=['BestBidQty', 'BestBid', 'BestAsk', 'BestAskQty', 'Interval'])
                books.loc[:, 'BestBidQty'] = book.loc[:, 'contractItems'].apply(lambda c: c.get('bestBidQty')).values
                books.loc[:, 'BestBid'] = book.loc[:, 'contractItems'].apply(lambda c: c.get('bestBidPr')).values
                books.loc[:, 'BestAskQty'] = book.loc[:, 'contractItems'].apply(lambda c: c.get('bestAskQty')).values
                books.loc[:, 'BestAsk'] = book.loc[:, 'contractItems'].apply(lambda c: c.get('bestAskPr')).values
                books.loc[:, 'Interval'] = book.loc[:, 'contractItems'].apply(lambda c: c.get('interval')).values
                books.loc[:, 'Flow date'] = book.loc[:, 'contractItems'].apply(lambda c: c.get('flowDate')).values
                books.loc[:, 'Flow date'] = book.loc[:, 'contractItems'].apply(lambda c: c.get('flowDate')).values
                books.loc[:, 'timeresolution'] = book.loc[:, 'contractItems'].apply(
                    lambda c: c.get('timeresolution')).values

                date_filter = datetime.strptime(self.flow_date, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
                date_filter = date_filter + timedelta(days=1)
                date_filter = date_filter.strftime("%y%m%d")

                books = books.loc[books.loc[:, 'Flow date'] == date_filter, :]

                if resolution == 'PT15M':
                    books = books.loc[books.loc[:, 'timeresolution'] == 'QH', :]

                elif resolution == 'PT60M':
                    books = books.loc[books.loc[:, 'timeresolution'] == 'FH', :]

                order_book = pd.concat([order_book, books])

        return trades


def create_payload(pos_list: List[int], purpose_list: List[str], price_list: List[float], qty_list: List[float], area_code: List[str], unit_code: List[str], flow_date: str, granularity: str, message: str):
    """
    Creates the dictionary to be sent to the market via API
    """

    if not (len(pos_list) == len(purpose_list) == len(price_list) == len(qty_list)):
        raise ValueError("All input lists must have the same length")

    orders = []
    for pos, purpose, price, qty, area, unit in zip(pos_list, purpose_list, price_list, qty_list, area_code, unit_code):
        order = {
            "operationType": "NEW",
            "areaCode": area,
            "unitCode": unit,
            "flowDate": flow_date,
            "pos": pos,
            "resolution": granularity,
            "orderType": "STD",
            "mode": "NON",
            "executionType": "NOR",
            "purpose": purpose.upper(),
            "price": price,
            "qty": qty
        }
        orders.append(order)

    payload = {
        "orderType": "BSK",
        "orderExecutionType": "NON",
        "hermes_txt": message,
        "orders": orders
    }

    return payload

def weighted_avg_price(group):
    return (group["price"] * group["quantity"]).sum() / group["quantity"].sum()

def find_closing_prices(df1, df2):
    # pivot df1 into BUy and SELL columns
    pivot = df1.pivot(index="period", columns="type", values="weighted_avg_price")

    # merge with exposure
    merged = pivot.merge(df2, on="period", how="left").fillna(0)

    results = []
    for _, row in merged.iterrows():
        period = row["period"]
        exp = row["exposure"]
        pb = row.get("BUY", None)  # avg buy price
        ps = row.get("SELL", None)  # avg sell price

        # assume unit volumes for existing trades
        qb = 1 if not pd.isna(pb) else 0
        qs = 1 if not pd.isna(ps) else 0

        if exp < 0:  # long -> must SELL
            pbuy = pb
            if qs == 0:
                x = pbuy + 4
                p_sell_new = x
            else:
                x = (((pbuy + 4) * (qs + abs(exp))) - (qs * ps)) / abs(exp)
                p_sell_new = (qs * ps + abs(exp) * x) / (qs + abs(exp))
            p_buy_new = pbuy
            results.append((period, "SELL", round(x, 2), round(p_buy_new, 2), round(p_sell_new, 2),
                            round(p_sell_new - p_buy_new, 2)))

        elif exp > 0:  # short -> must BUy
            psell = ps
            if qb == 0:
                y = psell - 4
                p_buy_new = y
            else:
                y = (((psell - 4) * (qb + exp)) - (qb * pb)) / exp
                p_buy_new = (qb * pb + exp * y) / (qb + exp)
            p_sell_new = psell
            results.append((period, "BUY", round(y, 2), round(p_buy_new, 2), round(p_sell_new, 2),
                            round(p_sell_new - p_buy_new, 2)))

    return pd.DataFrame(results, columns=["period","closing_type","closing_price","new_buy_avg","new_sell_avg","spread"])



def create_auction_payload(pos_list: List[int], purpose_list: List[str], price_list: List[float], qty_list: List[float], area_code: List[str], unit_code: List[str], flow_date: str, granularity: str, market: str):
    """
    Creates the dictionary to be sent to the market via API
    """

    if not (len(pos_list) == len(purpose_list) == len(price_list) == len(qty_list)):
        raise ValueError("All input lists must have the same length")

    orders = []
    for pos, purpose, price, qty, area, unit in zip(pos_list, purpose_list, price_list, qty_list, area_code, unit_code):
        order = {
            "unit_reference_number": unit,
            "date": flow_date,
            "pos": pos,
            "predefined_offer": False,
            "resolution": granularity,
            "purpose": purpose,
            "energy_price": price,
            "bid_quantity": qty,
            "market": market
        }
        orders.append(order)

    payload = {
        "company_name": "enet",
        "offers": orders
    }

    return payload
