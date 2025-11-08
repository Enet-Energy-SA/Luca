import pandas as pd
import openpyxl
import xlwings as xw
from datetime import datetime

mapping = {'NORD': 1, 'CNOR': 2, 'CSUD': 3, 'SUD': 4, 'CALA': 5, 'SICI': 6, 'SARD': 7}

df_prezzi = pd.DataFrame()

for zona in ['NORD', 'CNOR', 'CSUD', 'SUD', 'SICI', 'CALA', 'SARD']:
    df_temp = pd.read_csv('C:\\Users\\lbellomi\\Downloads\\PricesTable (' + str(mapping[zona]) + ').csv', sep=';')
    df_temp = df_temp.loc[:, ['Date', 'Year', 'Month', 'Day', 'Period', 'PUN', 'MGP', 'MI1', 'MI2', 'MI3', 'IMB']]
    df_temp.loc[:, 'zona'] = zona

    df_prezzi = pd.concat([df_prezzi, df_temp])

df_prezzi.loc[:, 'Date'] = df_prezzi.loc[:, 'Date'].apply(lambda d: datetime.strptime(d, "%Y-%m-%d").date())

file = r"C:\Users\lbellomi\PycharmProjects\pythonProject\Trader\bids.xlsm"
wb = openpyxl.load_workbook(file, keep_vba=True)
# Select sheet "prices"
ws = wb["Send Bids"]
Data = ws["B2"].value
ws = wb["Prices"]
zone = [cell.value for cell in ws["K"]]
#zone = ['NORD', 'CSUD', 'SUD', 'SICI']
zone = ['NORD', 'CSUD', 'SUD', 'SICI', 'CNOR', 'CALA', 'SARD']

df_prezzi = df_prezzi.loc[df_prezzi.loc[:, 'Date'] == Data.date(),:]
df_prezzi = df_prezzi.loc[df_prezzi.loc[:, 'zona'].isin(zone),:]

start_row = 2
mapping_mi = {'NORD': 2, 'CSUD': 3, 'SUD': 4, 'SICI': 5, 'CNOR': 7, 'CALA':6, 'SARD': 8}
mapping_mgp = {'NORD': 12, 'CSUD': 13, 'SUD': 14, 'SICI': 15, 'CNOR': 17, 'CALA': 16, 'SARD': 18}

for zona in zone:

    P = df_prezzi.loc[df_prezzi.loc[:, 'zona'] == zona, ['Period', 'MI1', 'MGP']]
    P = P.sort_values(by='Period')

    for i, (mi, mgp) in enumerate(zip(P.loc[:, 'MI1'], P.loc[:, 'MGP']), start=start_row):
        ws.cell(row=i, column=mapping_mi[zona], value=mi)
        ws.cell(row=i, column=mapping_mgp[zona], value=mgp)
        ws.cell(row=i, column=22, value=Data)

wb.save(file)

